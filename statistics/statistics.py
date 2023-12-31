from string import ascii_uppercase
from typing import Dict, Tuple

from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import PatternFill

from parsing.models import MonthCategories
from settings import (COLOR_BAD, COLOR_GOOD, INVERT_MONTHS, PATH_TO_DATA,
                      RESULT_NAME_SHEET, COLOR_NEUTRAL)


def get_statistics(data, expenditure_category: Dict[str, int], income_category: Dict[str, int]) -> None:
    """Главная функция создания статистики"""
    workbook = load_workbook(PATH_TO_DATA)
    if RESULT_NAME_SHEET in workbook.sheetnames:
        del workbook[RESULT_NAME_SHEET]
    worksheet = workbook.create_sheet(RESULT_NAME_SHEET)

    worksheet["A1"] = "Баланс:"

    global_expenditure = 0
    global_income = 0
    index_number = 4

    for month_category in data:
        month_category: MonthCategories
        current_expenditure, current_income = __create_month_description(worksheet, month_category, index_number)
        __create_charts(worksheet, index_number, month_category, expenditure_category, income_category)

        after = global_income - global_expenditure
        global_expenditure += current_expenditure
        global_income += current_income

        __create_month_description2(worksheet, after, global_income - global_expenditure, index_number)

        index_number += 10

    worksheet["B1"] = global_income - global_expenditure
    workbook.save(PATH_TO_DATA)


def __create_month_description2(worksheet, after: int, before: int, index_number: int) -> None:
    """Создание полей: сумма до начала месяца и после"""
    worksheet[f"B{index_number + 2}"] = "До"
    worksheet[f"C{index_number + 2}"] = after

    worksheet[f"B{index_number + 3}"] = "После"
    worksheet[f"C{index_number + 3}"] = before


def __create_month_description(worksheet, month_category: MonthCategories, index_number: int) -> Tuple[int, int]:
    """Создание полей: месяц, расход, доход, дельта"""
    month_category: MonthCategories
    worksheet[f"A{index_number + 1}"] = INVERT_MONTHS[month_category.month]
    expenditure, income = month_category.get_expenditure(), month_category.get_income()

    worksheet[f"B{index_number - 1}"] = "Расход"
    worksheet[f"C{index_number - 1}"] = expenditure

    worksheet[f"B{index_number}"] = "Доход"
    worksheet[f"C{index_number}"] = income

    worksheet[f"B{index_number + 1}"] = "Дельта"
    worksheet[f"C{index_number + 1}"] = income - expenditure

    __create_font(worksheet, index_number)

    return expenditure, income


def __create_font(worksheet, index_number: int) -> None:
    """Создание подцветки для описание"""
    worksheet[f"B{index_number - 1}"].fill = worksheet[f"C{index_number - 1}"].fill =\
        PatternFill(start_color=COLOR_BAD, end_color=COLOR_BAD, fill_type="solid")

    worksheet[f"B{index_number}"].fill = worksheet[f"C{index_number}"].fill =\
        PatternFill(start_color=COLOR_GOOD, end_color=COLOR_GOOD, fill_type="solid")

    if worksheet[f"C{index_number + 1}"].value >= 0:
        worksheet[f"C{index_number + 1}"].fill = worksheet[f"B{index_number + 1}"].fill =\
            PatternFill(start_color=COLOR_GOOD, end_color=COLOR_GOOD, fill_type="solid")
    else:
        worksheet[f"C{index_number + 1}"].fill = worksheet[f"B{index_number + 1}"].fill =\
            PatternFill(start_color=COLOR_BAD, end_color=COLOR_BAD, fill_type="solid")

    worksheet[f"B{index_number + 2}"].fill = worksheet[f"C{index_number + 2}"].fill = \
        PatternFill(start_color=COLOR_NEUTRAL, end_color=COLOR_NEUTRAL, fill_type="solid")

    worksheet[f"B{index_number + 3}"].fill = worksheet[f"C{index_number + 3}"].fill = \
        PatternFill(start_color=COLOR_NEUTRAL, end_color=COLOR_NEUTRAL, fill_type="solid")


def __create_charts(worksheet, index_number: int, month_category: MonthCategories,
                    expenditure_category: Dict[str, int], income_category: Dict[str, int]) -> None:
    """Создание графиков, запись данных для графиков"""
    max_col_income = __write_data_for_chart(worksheet, index_number, month_category.income_data)
    max_col_expenditure = __write_data_for_chart(worksheet, index_number+2, month_category.expenditure_data)

    __draw_pie_chart(worksheet=worksheet, index_number=index_number, max_col=max_col_income,
                     title=f"Доходы {INVERT_MONTHS[month_category.month]}", chart_index=f"E{index_number-1}",
                     data=month_category.income_data, category=income_category)
    __draw_pie_chart(worksheet=worksheet, index_number=index_number+2, max_col=max_col_expenditure,
                     title=f"Расходы {INVERT_MONTHS[month_category.month]}", chart_index=f"K{index_number-1}",
                     data=month_category.expenditure_data, category=expenditure_category)


def __write_data_for_chart(worksheet, index_number: int, data: Dict[str, int]) -> int:
    """Запись значений для графиков"""
    row_index = 10
    for key, value in data.items():
        worksheet[f"{ascii_uppercase[row_index]}{index_number-1}"] = key
        worksheet[f"{ascii_uppercase[row_index]}{index_number}"] = value
        row_index += 1
    return row_index


def setup_pie_chart_color(pie, data, labels, data_ref, category) -> None:
    """Установка цвета на графике под каждую категорию"""
    series = Series(data_ref)

    for i, item in enumerate(data.items()):
        key, value = item
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = category[key]
        series.dPt.append(pt)

    pie.append(series)
    pie.set_categories(labels)


def __draw_pie_chart(worksheet, index_number: int, max_col: int, title: str, chart_index: str, data, category):
    """Рисование графиков"""
    pie = PieChart()

    labels = Reference(worksheet, min_col=11, max_col=max_col, min_row=index_number-1, max_row=index_number-1)
    data_ref = Reference(worksheet, min_col=11, max_col=max_col, min_row=index_number, max_row=index_number)

    setup_pie_chart_color(pie, data, labels, data_ref, category)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = True

    pie.title = title
    pie.height = 5
    pie.width = 10

    worksheet.add_chart(pie, chart_index)
